from __future__ import annotations

import calendar
import json
import math
import os
import re
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook


SHEET_ALIASES = {
    "booked": ["Shipment Booked Yesterday", "Booked Yesterday", "Shipment Booked"],
    "daily_tonnage": ["Daily Tonnage"],
    "month_tonnage": ["Total_Tonnage_this_month", "Total Tonnage this month", "Month Tonnage"],
    "target_plan": ["Tonnage_of_June_Month", "Tonnage", "Target", "Monthly Target"],
    "due": ["Order due Tommorow", "Order due Tomorrow", "Due Tomorrow"],
    "edd": ["Order EDD Crossed", "EDD Crossed"],
    "open": ["Open Shipment", "Open Shipments"],
}

# Normalized raw names -> normalized target/client names. This handles short names and spelling
# variants found in operational exports while keeping the display names from the target sheet.
NAME_ALIASES = {
    "onericappliances": "oneiricappliances",
    "oneiricappliances": "oneiricappliances",
    "mitras": "mitrastechnocrafts",
    "mitrastechnocraft": "mitrastechnocrafts",
    "sukuga": "sukugatechnologies",
    "sukugatechnology": "sukugatechnologies",
    "carrierctd": "carrierctd",
    "carrierctdglobal": "carrierctd",
}

STOP_WORDS = {
    "pvt",
    "private",
    "ltd",
    "limited",
    "llp",
    "company",
    "co",
    "hr",
}


@dataclass
class SourceResult:
    excel_path: Path
    temporary: bool = False


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return str(value).strip() == ""


def _as_date(value: Any) -> Optional[date]:
    if _is_blank(value):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip()
    if not text or text.lower() in {"nat", "nan", "none"}:
        return None

    # Common ISO/date formats exported by Google Sheets and Excel.
    candidates = [
        text,
        text.replace("/", "-"),
        text.split(" ")[0].replace("/", "-"),
        text.split("T")[0].replace("/", "-"),
    ]
    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%m-%d-%Y %H:%M:%S",
    ]
    for cand in dict.fromkeys(candidates):
        try:
            return datetime.fromisoformat(cand).date()
        except Exception:
            pass
        for fmt in formats:
            try:
                return datetime.strptime(cand, fmt).date()
            except Exception:
                pass
    return None


def _is_number(value: Any) -> bool:
    if isinstance(value, bool) or _is_blank(value):
        return False
    if isinstance(value, (int, float)):
        return not (isinstance(value, float) and math.isnan(value))
    try:
        float(str(value).replace(",", ""))
        return True
    except Exception:
        return False


def _num(value: Any, default: float = 0.0) -> float:
    if not _is_number(value):
        return default
    return float(str(value).replace(",", ""))


def normalize_name(name: Any) -> str:
    """Return a stable comparison key for customer names."""
    if _is_blank(name):
        return ""
    s = str(name).lower().strip().replace("&", "and")
    tokens = re.findall(r"[a-z0-9]+", s)
    tokens = [t for t in tokens if t not in STOP_WORDS]
    key = "".join(tokens)
    return NAME_ALIASES.get(key, key)


def clean_display_name(name: Any) -> str:
    return re.sub(r"\s+", " ", str(name).strip())


class NameResolver:
    def __init__(self) -> None:
        self.display_by_key: dict[str, str] = {}

    def add(self, display: str) -> str:
        display = clean_display_name(display)
        key = normalize_name(display)
        if key and key not in self.display_by_key:
            self.display_by_key[key] = display
        return display

    def resolve(self, raw: Any) -> Optional[str]:
        if _is_blank(raw):
            return None
        display = clean_display_name(raw)
        if not display or display.lower() in {"nan", "nat", "none", "total", "grand total"}:
            return None
        key = normalize_name(display)
        if not key:
            return None
        if key in self.display_by_key:
            return self.display_by_key[key]
        # Short operational names such as "Loom Solar" should match target names like
        # "Loom Solar Pvt Ltd".
        for target_key, target_display in self.display_by_key.items():
            if key and (key in target_key or target_key in key):
                return target_display
        self.display_by_key[key] = display
        return display


def google_sheet_export_url(url: str) -> str:
    """Convert a normal Google Sheets URL to an XLSX export URL."""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not m:
        return url
    sheet_id = m.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def prepare_source(source: str | Path) -> SourceResult:
    """Download a URL or return a local Excel file path."""
    source_str = str(source)
    if re.match(r"https?://", source_str):
        export_url = google_sheet_export_url(source_str)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_path = Path(tmp.name)
        tmp.close()
        resp = requests.get(export_url, timeout=90)
        resp.raise_for_status()
        tmp_path.write_bytes(resp.content)
        return SourceResult(tmp_path, temporary=True)

    path = Path(source_str).expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"Excel source not found: {path}")
    return SourceResult(path, temporary=False)


def _norm_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def _sheet_rows(wb: Any, sheet_name: str) -> list[list[Any]]:
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _cell(rows: list[list[Any]], r: int, c: int) -> Any:
    if r < 0 or r >= len(rows):
        return None
    row = rows[r]
    if c < 0 or c >= len(row):
        return None
    return row[c]


def find_sheet(wb: Any, kind: str) -> Optional[str]:
    wanted = [_norm_text(x) for x in SHEET_ALIASES.get(kind, [])]
    actual = {_norm_text(s): s for s in wb.sheetnames}
    for w in wanted:
        if w in actual:
            return actual[w]
    for key, name in actual.items():
        if any(w in key or key in w for w in wanted):
            return name
    return None


def find_column_index(headers: list[Any], candidates: list[str]) -> Optional[int]:
    normalized = {_norm_text(c): idx for idx, c in enumerate(headers)}
    for cand in candidates:
        key = _norm_text(cand)
        if key in normalized:
            return normalized[key]
    for idx, header in enumerate(headers):
        h = _norm_text(header)
        if any(_norm_text(cand) in h for cand in candidates):
            return idx
    return None


def _find_header_row_for_columns(rows: list[list[Any]], candidates: list[str], max_scan: int = 15) -> Optional[int]:
    for r in range(min(max_scan, len(rows))):
        if find_column_index(rows[r], candidates) is not None:
            return r
    return None


def _mode_date(values: list[Any]) -> Optional[date]:
    dates = [d for d in (_as_date(v) for v in values) if d]
    if not dates:
        return None
    return Counter(dates).most_common(1)[0][0]


def _find_target_header_row(rows: list[list[Any]]) -> Optional[int]:
    for idx in range(min(10, len(rows))):
        values = " | ".join(str(v).lower() for v in rows[idx])
        if "name of the personnel" in values and "name of the customer" in values:
            return idx
    return None


def _find_day_start_col(rows: list[list[Any]], header_row: int) -> int:
    header = rows[header_row]
    for col, value in enumerate(header):
        txt = str(value).lower()
        if "monthly" in txt and "target" in txt:
            return min(col + 1, len(header))
    return 5


def infer_target_last_updated_day(wb: Any) -> Optional[int]:
    """Return the last day number updated in the KAM/person tonnage sheet.

    This helper is kept for reference/debugging, but active days currently follow
    the business rule in current_date_minus_one_active_days().
    """
    target_sheet = find_sheet(wb, "target_plan")
    if not target_sheet:
        return None

    rows = _sheet_rows(wb, target_sheet)
    header_row = _find_target_header_row(rows)
    if header_row is None:
        return None

    day_start = _find_day_start_col(rows, header_row)
    header_len = len(rows[header_row])
    last_day = 0
    for col in range(day_start, header_len):
        if any(_is_number(_cell(rows, r, col)) for r in range(header_row + 1, len(rows))):
            last_day = col - day_start + 1
    return last_day or None


def current_date_minus_one_active_days() -> int:
    """Active days = current calendar date - 1, including Sundays/holidays.

    Example: if the backend runs on 22 June, active days = 21 for every client,
    even if 21 June was Sunday. Timezone defaults to Asia/Kolkata and can be
    changed with MVIKAS_TIMEZONE.
    """
    timezone_name = os.getenv("MVIKAS_TIMEZONE", "Asia/Kolkata")
    try:
        today = datetime.now(ZoneInfo(timezone_name)).date()
    except Exception:
        today = datetime.now().date()
    return max(today.day - 1, 0)


def infer_report_date(wb: Any) -> date:
    """Infer dashboard 'as-of' date from due/booked sheets; fallback to target sheet or today."""
    due_sheet = find_sheet(wb, "due")
    if due_sheet:
        rows = _sheet_rows(wb, due_sheet)
        header_row = _find_header_row_for_columns(rows, ["EDD", "ExpectedDelivery", "Expected Delivery", "v_expected_delivery_date"])
        if header_row is not None:
            col = find_column_index(rows[header_row], ["EDD", "ExpectedDelivery", "Expected Delivery", "v_expected_delivery_date"])
            if col is not None:
                d = _mode_date([_cell(rows, r, col) for r in range(header_row + 1, len(rows))])
                if d:
                    return d - timedelta(days=1)

    booked_sheet = find_sheet(wb, "booked")
    if booked_sheet:
        rows = _sheet_rows(wb, booked_sheet)
        header_row = _find_header_row_for_columns(rows, ["Booking_date", "Booking Date", "v_booking_date"])
        if header_row is not None:
            col = find_column_index(rows[header_row], ["Booking_date", "Booking Date", "v_booking_date"])
            if col is not None:
                d = _mode_date([_cell(rows, r, col) for r in range(header_row + 1, len(rows))])
                if d:
                    return d + timedelta(days=1)

    target_sheet = find_sheet(wb, "target_plan")
    if target_sheet:
        rows = _sheet_rows(wb, target_sheet)
        header_row = _find_target_header_row(rows)
        if header_row is not None:
            day_start = _find_day_start_col(rows, header_row)
            first_day = None
            for col in range(day_start, len(rows[header_row])):
                first_day = _as_date(_cell(rows, header_row, col))
                if first_day:
                    break
            last_idx = 0
            for col in range(day_start, len(rows[header_row])):
                if any(_is_number(_cell(rows, r, col)) for r in range(header_row + 1, len(rows))):
                    last_idx = max(last_idx, col - day_start + 1)
            if first_day and last_idx:
                days_in_month = calendar.monthrange(first_day.year, first_day.month)[1]
                return date(first_day.year, first_day.month, min(last_idx, days_in_month))

    return datetime.now().date()


def parse_target_plan(
    wb: Any,
    report_date: date,
    resolver: NameResolver,
    active_day_override: Optional[int] = None,
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]: 
    sheet = find_sheet(wb, "target_plan")
    clients: list[dict[str, Any]] = []
    by_key: dict[str, dict[str, Any]] = {}
    if not sheet:
        return clients, by_key

    rows = _sheet_rows(wb, sheet)
    header_row = _find_target_header_row(rows)
    if header_row is None:
        return clients, by_key

    header = rows[header_row]

    def col_contains(*words: str, default: int) -> int:
        for i, h in enumerate(header):
            h_text = str(h).lower()
            if all(w.lower() in h_text for w in words):
                return i
        return default

    person_col = col_contains("personnel", default=1)
    customer_col = col_contains("customer", default=3)
    target_col = col_contains("monthly", "target", default=4)
    day_start = _find_day_start_col(rows, header_row)
    active_days_for_all = active_day_override if active_day_override is not None else report_date.day
    max_day_cols = max(0, min(active_days_for_all, len(header) - day_start))

    for r in range(header_row + 1, len(rows)):
        raw_customer = _cell(rows, r, customer_col)
        if _is_blank(raw_customer):
            continue
        customer = clean_display_name(raw_customer)
        if not customer or customer.lower() in {"nan", "nat", "none"}:
            continue
        display = resolver.add(customer)
        key = normalize_name(display)

        person = clean_display_name(_cell(rows, r, person_col)) if not _is_blank(_cell(rows, r, person_col)) else "Not Alloted"
        if person.lower() in {"nan", "nat", "none", ""}:
            person = "Not Alloted"

        target = _num(_cell(rows, r, target_col), 0.0)
        day_values = [_cell(rows, r, c) for c in range(day_start, day_start + max_day_cols)]
        # Active days are now the sheet's last updated day for every client.
        # Example: if data is updated through 17 June, activeDays = 17 for all rows.
        active_days = max_day_cols
        plan_sum = round(sum(_num(v) for v in day_values if _is_number(v)), 2)

        client = {
            "name": display,
            "person": person,
            "target": round(target, 2),
            "achieved": plan_sum,
            "activeDays": active_days,
        }
        clients.append(client)
        by_key[key] = client

    return clients, by_key


def group_count(wb: Any, kind: str, resolver: NameResolver) -> list[dict[str, Any]]:
    sheet = find_sheet(wb, kind)
    if not sheet:
        return []
    rows = _sheet_rows(wb, sheet)
    header_row = _find_header_row_for_columns(rows, ["Customername", "Customer Name", "customername", "customer"])
    if header_row is None:
        return []
    ccol = find_column_index(rows[header_row], ["Customername", "Customer Name", "customername", "customer"])
    if ccol is None:
        return []

    counts: dict[str, int] = defaultdict(int)
    for r in range(header_row + 1, len(rows)):
        name = resolver.resolve(_cell(rows, r, ccol))
        if name:
            counts[name] += 1
    return [{"name": k, "count": int(v)} for k, v in sorted(counts.items(), key=lambda kv: kv[1], reverse=True)]


def group_tonnage(wb: Any, kind: str, resolver: NameResolver) -> list[dict[str, Any]]:
    sheet = find_sheet(wb, kind)
    if not sheet:
        return []
    rows = _sheet_rows(wb, sheet)
    header_row = _find_header_row_for_columns(rows, ["customername", "Customername", "Customer Name", "customer"])
    if header_row is None:
        return []
    ccol = find_column_index(rows[header_row], ["customername", "Customername", "Customer Name", "customer"])
    kg_col = find_column_index(rows[header_row], ["Total_Tonnage_in_Kg", "Total Tonnage in Kg", "Tonnage", "kg"])
    if ccol is None or kg_col is None:
        return []

    totals: dict[str, float] = defaultdict(float)
    for r in range(header_row + 1, len(rows)):
        name = resolver.resolve(_cell(rows, r, ccol))
        if not name:
            continue
        kg = _num(_cell(rows, r, kg_col), 0.0)
        if kg:
            totals[name] += kg
    field = "kg" if kind == "daily_tonnage" else "achieved"
    return [{"name": k, field: round(v, 2)} for k, v in sorted(totals.items(), key=lambda kv: kv[1], reverse=True)]


def extract_mode_date_from_sheet(wb: Any, kind: str, columns: list[str]) -> Optional[date]:
    sheet = find_sheet(wb, kind)
    if not sheet:
        return None
    rows = _sheet_rows(wb, sheet)
    header_row = _find_header_row_for_columns(rows, columns)
    if header_row is None:
        return None
    col = find_column_index(rows[header_row], columns)
    if col is None:
        return None
    return _mode_date([_cell(rows, r, col) for r in range(header_row + 1, len(rows))])


def _parse_report_date(value: str | date) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    parsed = _as_date(value)
    if not parsed:
        raise ValueError(f"Invalid report date: {value}")
    return parsed


def parse_excel(source: str | Path, report_date: str | date | None = None) -> dict[str, Any]:
    """Parse the MVIKAS workbook using openpyxl only.

    This intentionally does not use pandas, so the backend works on newer Windows Python
    versions where pandas wheels may not be available yet.
    """
    src = prepare_source(source)
    try:
        wb = load_workbook(src.excel_path, read_only=True, data_only=True)
        if report_date:
            as_of = _parse_report_date(report_date)
        else:
            as_of = infer_report_date(wb)

        resolver = NameResolver()
        # Business rule: active days are always current date - 1.
        # Example: if today is 22nd, activeDays = 21 for all clients,
        # even if yesterday was Sunday/holiday.
        active_days_for_all = current_date_minus_one_active_days()
        clients, clients_by_key = parse_target_plan(
            wb,
            as_of,
            resolver,
            active_day_override=active_days_for_all,
        )

        # Authoritative monthly achieved values override the sum in target-plan columns when available.
        monthly_rows = group_tonnage(wb, "month_tonnage", resolver)
        monthly_by_key = {normalize_name(r["name"]): float(r["achieved"]) for r in monthly_rows}
        for client in clients:
            key = normalize_name(client["name"])
            if key in monthly_by_key:
                client["achieved"] = round(monthly_by_key[key], 2)

        # Add monthly tonnage customers that are not present in the target sheet.
        for row in monthly_rows:
            key = normalize_name(row["name"])
            if key not in clients_by_key:
                client = {
                    "name": resolver.add(row["name"]),
                    "person": "Not Alloted",
                    "target": 0.0,
                    "achieved": float(row["achieved"]),
                    "activeDays": active_days_for_all,
                }
                clients.append(client)
                clients_by_key[key] = client

        clients.sort(key=lambda c: (float(c.get("achieved", 0)), float(c.get("target", 0))), reverse=True)

        open_data = group_count(wb, "open", resolver)
        edd_data = group_count(wb, "edd", resolver)
        due_data = group_count(wb, "due", resolver)
        booked_data = group_count(wb, "booked", resolver)
        daily_tonnage = group_tonnage(wb, "daily_tonnage", resolver)

        due_date = extract_mode_date_from_sheet(wb, "due", ["EDD", "ExpectedDelivery", "Expected Delivery", "v_expected_delivery_date"]) or (as_of + timedelta(days=1))
        booked_date = extract_mode_date_from_sheet(wb, "booked", ["Booking_date", "Booking Date", "v_booking_date"]) or (as_of - timedelta(days=1))
        daily_tonnage_date = booked_date

        days_in_month = calendar.monthrange(as_of.year, as_of.month)[1]
        # Dashboard active days / daily average denominator follows current date - 1.
        # If today is 22nd, this is 21, including Sundays/holidays.
        active_days_elapsed = min(max(active_days_for_all, 1), days_in_month)

        return {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sourceFile": str(src.excel_path.name),
            "reportDate": as_of.isoformat(),
            "bookedDate": booked_date.isoformat(),
            "dueDate": due_date.isoformat(),
            "dailyTonnageDate": daily_tonnage_date.isoformat(),
            "monthName": as_of.strftime("%B"),
            "monthYear": as_of.strftime("%B %Y"),
            "year": as_of.year,
            "month": as_of.month,
            "daysInMonth": days_in_month,
            "activeDaysElapsed": active_days_elapsed,
            "clients": clients,
            "eddData": edd_data,
            "openData": open_data,
            "dueData": due_data,
            "bookedData": booked_data,
            "dailyTonnageData": daily_tonnage,
        }
    finally:
        if src.temporary:
            try:
                src.excel_path.unlink(missing_ok=True)
            except Exception:
                pass


def write_json(data: dict[str, Any], path: str | Path) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
